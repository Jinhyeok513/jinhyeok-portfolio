from __future__ import annotations

import csv
import shutil
import zipfile
from collections import Counter, defaultdict, deque
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path("/Users/jin/Desktop/Event Management System Project Plan")
PROJECT = ROOT / "projects" / "ems-project-management-case-study"
ORIGINAL = PROJECT / "original-artifacts"
CORRECTED = PROJECT / "corrected"
ASSETS = PROJECT / "assets"
PUBLIC_IMAGE = ROOT / "public" / "images" / "projects" / "ems-gantt-preview.png"
PUBLIC_CASE_STUDY = ROOT / "public" / "case-studies" / "ems.md"


SOURCE_FILES = {
    "AON Diagram copy.png": "AON_Diagram_original.png",
    "EMS_AON_Full_Detail.pdf": "EMS_AON_Full_Detail.pdf",
    "Gantt Chart  copy.xlsx": "Gantt_Chart_original.xlsx",
    "WBS_List  copy.xlsx": "WBS_List_original.xlsx",
    "Project report.docx": "Project_Report_original.docx",
}


@dataclass(frozen=True)
class TaskDef:
    wbs: str
    title: str
    duration: int
    predecessors: tuple[str, ...]
    phase: str
    workstream: str
    owner: str
    deliverable: str


TASKS = [
    TaskDef("1.1.1", "Confirm NEC/RAH objectives and EMS success criteria", 5, (), "Initiation and planning", "Governance", "Project Director", "Approved business objectives and MOV"),
    TaskDef("1.1.2", "Map stakeholders and communication needs", 7, ("1.1.1",), "Initiation and planning", "Stakeholder management", "Project Director", "Stakeholder register and communication matrix"),
    TaskDef("1.1.3", "Define EMS project charter", 8, ("1.1.2",), "Initiation and planning", "Governance", "Project Director", "Project charter"),
    TaskDef("1.2.1", "Create scope baseline for registration, ticketing, schedule, budget, and security modules", 10, ("1.1.3",), "Initiation and planning", "Scope", "System Architect", "Scope statement and exclusions"),
    TaskDef("1.2.2", "Develop WBS dictionary and acceptance criteria", 7, ("1.2.1",), "Initiation and planning", "Scope", "Scrum Master / Agile Coach", "WBS dictionary"),
    TaskDef("1.3.1", "Build risk, quality, procurement, and change-control plans", 12, ("1.2.2",), "Initiation and planning", "Controls", "Operations Manager", "Integrated management control plan"),
    TaskDef("1.4.1", "Baseline budget and resource plan within AUD 1.5M cap", 8, ("1.3.1",), "Initiation and planning", "Cost and resources", "Project Director", "Budget and resource baseline"),
    TaskDef("1.5.1", "Milestone: planning baseline approved", 0, ("1.4.1",), "Initiation and planning", "Milestone", "PMP Steering Group", "Planning sign-off"),
    TaskDef("2.1.1", "Define EMS solution architecture and integration map", 15, ("1.5.1",), "Design and development", "Architecture", "System Architect", "Architecture baseline"),
    TaskDef("2.1.2", "Design data model, privacy controls, and cybersecurity architecture", 12, ("2.1.1",), "Design and development", "Architecture", "System Architect", "Security and data design"),
    TaskDef("2.1.3", "Prepare UI/UX prototypes for attendee, vendor, staff, and NEC admin flows", 20, ("2.1.1",), "Design and development", "Experience design", "UI/UX Designer", "Prototype set"),
    TaskDef("2.2.1", "Build online registration module", 35, ("2.1.2", "2.1.3"), "Design and development", "Core platform", "Frontend Developer", "Registration workflow"),
    TaskDef("2.2.2", "Build ticketing, payment, QR/RFID validation, and entry scanning module", 70, ("2.2.1",), "Design and development", "Core platform", "Ticketing System Specialist", "Ticketing and access-control workflow"),
    TaskDef("2.2.3", "Build performance scheduling and rehearsal management module", 45, ("2.1.2", "2.1.3"), "Design and development", "Core platform", "Backend Developer", "Schedule management workflow"),
    TaskDef("2.2.4", "Build budget analytics dashboard and reporting views", 40, ("2.1.2",), "Design and development", "Analytics", "Database Administrator", "Budget analytics module"),
    TaskDef("2.2.5", "Build security management dashboard and incident reporting workflow", 55, ("2.1.2",), "Design and development", "Security", "Cybersecurity Engineer", "Security dashboard"),
    TaskDef("2.3.1", "Integrate vendor APIs, mobile app views, and NEC admin portal", 40, ("2.2.2", "2.2.3", "2.2.4", "2.2.5"), "Design and development", "Integration", "API Engineer", "Integrated EMS platform"),
    TaskDef("2.4.1", "Complete internal sprint reviews and defect triage", 20, ("2.3.1",), "Design and development", "Quality", "Scrum Master / Agile Coach", "Sprint review pack"),
    TaskDef("2.5.1", "Milestone: integrated EMS build complete", 0, ("2.4.1",), "Design and development", "Milestone", "PMP Steering Group", "Build completion sign-off"),
    TaskDef("3.1.1", "Run system integration testing across all EMS modules", 35, ("2.5.1",), "Testing and validation", "Testing", "QA Test Engineer", "SIT report"),
    TaskDef("3.1.2", "Run security, privacy, and penetration testing", 25, ("3.1.1",), "Testing and validation", "Security testing", "Security Tester", "Security test report"),
    TaskDef("3.1.3", "Run performance, load, and entry-scanning stress testing", 20, ("3.1.1",), "Testing and validation", "Performance testing", "Performance Test Engineer", "Load-test report"),
    TaskDef("3.2.1", "Conduct NEC staff UAT and festival operations rehearsal", 25, ("3.1.2", "3.1.3"), "Testing and validation", "User acceptance", "Operations Manager", "UAT sign-off log"),
    TaskDef("3.2.2", "Resolve priority defects and update training materials", 20, ("3.2.1",), "Testing and validation", "Defect remediation", "Scrum Master / Agile Coach", "Defect closure and training pack"),
    TaskDef("3.3.1", "Milestone: EMS validated for deployment", 0, ("3.2.2",), "Testing and validation", "Milestone", "PMP Steering Group", "Deployment approval"),
    TaskDef("4.1.1", "Prepare cloud production environment and monitoring", 20, ("3.3.1",), "Deployment and readiness", "Release", "DevOps Engineer", "Production environment"),
    TaskDef("4.1.2", "Migrate configuration, event data, ticket templates, and vendor records", 15, ("4.1.1",), "Deployment and readiness", "Release", "Database Administrator", "Migration checklist"),
    TaskDef("4.2.1", "Train NEC operations, security, ticketing, and vendor-support users", 25, ("4.1.2",), "Deployment and readiness", "Training", "Operations Manager", "Training completion register"),
    TaskDef("4.3.1", "Run production pilot and go-live rehearsal", 20, ("4.2.1",), "Deployment and readiness", "Release", "Project Director", "Pilot readiness report"),
    TaskDef("4.4.1", "Operate final contingency and defect-remediation window", 45, ("4.3.1",), "Deployment and readiness", "Controls", "PMP Delivery Team", "Readiness reserve log"),
    TaskDef("4.5.1", "Milestone: EMS ready for 2027 Sydney Festival operations", 0, ("4.4.1",), "Deployment and readiness", "Milestone", "NEC Sponsor", "Operational readiness sign-off"),
    TaskDef("5.1.1", "Complete handover, lessons learned, and support transition", 10, ("4.5.1",), "Closeout", "Handover", "Project Director", "Handover pack"),
    TaskDef("5.2.1", "Archive project evidence and final management report", 6, ("5.1.1",), "Closeout", "Reporting", "Scrum Master / Agile Coach", "Final report archive"),
    TaskDef("5.3.1", "Milestone: PMP final submission complete", 0, ("5.2.1",), "Closeout", "Milestone", "PMP Steering Group", "Final submission sign-off"),
]


PHASE_SUMMARY = [
    ("1", "Initiation and planning", "Define governance, scope, stakeholders, controls, budget, and delivery baseline."),
    ("2", "Design and development", "Design and build the integrated EMS platform modules."),
    ("3", "Testing and validation", "Validate the EMS through system, security, performance, and UAT activities."),
    ("4", "Deployment and readiness", "Prepare production, migrate configuration, train users, and confirm operational readiness."),
    ("5", "Closeout", "Handover the project and archive final evidence."),
]


def ensure_dirs() -> None:
    for path in (PROJECT, ORIGINAL, CORRECTED, ASSETS, PUBLIC_IMAGE.parent, PUBLIC_CASE_STUDY.parent):
        path.mkdir(parents=True, exist_ok=True)


def copy_originals() -> None:
    for source_name, dest_name in SOURCE_FILES.items():
        shutil.copy2(SOURCE / source_name, ORIGINAL / dest_name)


def parse_csv_refs(value: object) -> list[str]:
    if value is None:
        return []
    text = str(value).replace("，", ",").strip()
    if text in {"", "-"}:
        return []
    return [part.strip() for part in text.split(",") if part.strip() and part.strip() != "-"]


def read_original_gantt() -> tuple[list[dict], dict]:
    wb = openpyxl.load_workbook(SOURCE / "Gantt Chart  copy.xlsx", data_only=True)
    ws = wb["Sheet1"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        rows.append(
            {
                "wbs": str(row[0]).strip(),
                "title": str(row[1]).strip() if row[1] else "",
                "duration": int(row[2]) if row[2] is not None else None,
                "predecessors": parse_csv_refs(row[3]),
                "successors": parse_csv_refs(row[4]),
                "start": row[5],
                "end": row[6],
                "float": row[11],
                "critical": row[12],
            }
        )

    ids = {row["wbs"] for row in rows}
    title_counts = Counter(row["title"] for row in rows if row["title"])
    issues = {
        "task_count": len(rows),
        "start": min(row["start"] for row in rows if isinstance(row["start"], datetime)).date().isoformat(),
        "finish": max(row["end"] for row in rows if isinstance(row["end"], datetime)).date().isoformat(),
        "finish_after_required_deadline": "2026-11-30",
        "positive_duration_milestones": [
            {"wbs": row["wbs"], "title": row["title"], "duration": row["duration"]}
            for row in rows
            if "[Milestone]" in row["title"] and row["duration"]
        ],
        "duplicate_task_titles": {title: count for title, count in title_counts.items() if count > 1},
        "unknown_predecessor_refs": sorted({pred for row in rows for pred in row["predecessors"] if pred not in ids}),
        "unknown_successor_refs": sorted({succ for row in rows for succ in row["successors"] if succ not in ids}),
    }
    return rows, issues


def recover_original_wbs_rows() -> list[list[object]]:
    xlsx = SOURCE / "WBS_List  copy.xlsx"
    ns = {"main": "http://purl.oclc.org/ooxml/spreadsheetml/main"}
    with zipfile.ZipFile(xlsx) as zf:
        shared_xml = ET.fromstring(zf.read("xl/sharedStrings.xml"))
        shared = []
        for si in shared_xml.findall("main:si", ns):
            text = "".join(t.text or "" for t in si.findall(".//main:t", ns))
            shared.append(text)
        sheet = ET.fromstring(zf.read("xl/worksheets/sheet1.xml"))

    rows: list[list[object]] = []
    for row in sheet.findall(".//main:row", ns):
        values: dict[int, object] = {}
        for cell in row.findall("main:c", ns):
            ref = cell.attrib.get("r", "")
            col_letters = "".join(ch for ch in ref if ch.isalpha())
            col_num = 0
            for ch in col_letters:
                col_num = col_num * 26 + ord(ch.upper()) - 64
            value = cell.find("main:v", ns)
            if value is None:
                continue
            raw = value.text or ""
            if cell.attrib.get("t") == "s":
                values[col_num] = shared[int(raw)]
            else:
                try:
                    num = float(raw)
                    values[col_num] = int(num) if num.is_integer() else num
                except ValueError:
                    values[col_num] = raw
        if values:
            rows.append([values.get(col) for col in range(1, 13)])
    return rows


def compute_schedule() -> list[dict]:
    task_map = {task.wbs: task for task in TASKS}
    indegree = {task.wbs: 0 for task in TASKS}
    children: dict[str, list[str]] = defaultdict(list)
    for task in TASKS:
        for pred in task.predecessors:
            children[pred].append(task.wbs)
            indegree[task.wbs] += 1

    queue = deque([task.wbs for task in TASKS if indegree[task.wbs] == 0])
    ordered: list[str] = []
    while queue:
        current = queue.popleft()
        ordered.append(current)
        for child in children[current]:
            indegree[child] -= 1
            if indegree[child] == 0:
                queue.append(child)
    if len(ordered) != len(TASKS):
        raise RuntimeError("Corrected schedule contains a dependency cycle.")

    start_anchor = datetime(2025, 7, 15)
    starts: dict[str, datetime] = {}
    ends: dict[str, datetime] = {}
    for wbs in ordered:
        task = task_map[wbs]
        start = max((ends[pred] for pred in task.predecessors), default=start_anchor)
        end = start if task.duration == 0 else start + timedelta(days=task.duration)
        starts[wbs] = start
        ends[wbs] = end

    successors = defaultdict(list)
    for task in TASKS:
        for pred in task.predecessors:
            successors[pred].append(task.wbs)

    rows = []
    for index, task in enumerate(TASKS, start=1):
        rows.append(
            {
                "Task ID": f"EMS-{index:03d}" if task.duration else f"M-{index:03d}",
                "WBS": task.wbs,
                "Task": task.title,
                "Phase": task.phase,
                "Workstream": task.workstream,
                "Owner": task.owner,
                "Duration (days)": task.duration,
                "Start": starts[task.wbs].date().isoformat(),
                "Finish": ends[task.wbs].date().isoformat(),
                "Predecessors": ", ".join(task.predecessors) or "-",
                "Successors": ", ".join(successors[task.wbs]) or "-",
                "Deliverable": task.deliverable,
                "Milestone": "Yes" if task.duration == 0 else "No",
            }
        )
    return rows


def write_csv(path: Path, rows: list[dict] | list[list[object]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        if rows and isinstance(rows[0], dict):
            writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
        else:
            writer = csv.writer(handle)
            writer.writerows(rows)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="16324F")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="E8F1F8")
    thin = Side(style="thin", color="D9E2EC")
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if len(row) > 3 and row[3].value in {"Initiation and planning", "Design and development", "Testing and validation", "Deployment and readiness", "Closeout"}:
            row[3].fill = section_fill

    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        max_len = max(len(str(ws.cell(row=row, column=column).value or "")) for row in range(1, min(ws.max_row, 80) + 1))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 42)
    ws.row_dimensions[1].height = 34


def write_corrected_workbook(schedule_rows: list[dict], original_issues: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Corrected Schedule"
    headers = list(schedule_rows[0].keys())
    ws.append(headers)
    for row in schedule_rows:
        ws.append([row[header] for header in headers])
    style_sheet(ws)

    wbs = wb.create_sheet("WBS Summary")
    wbs.append(["WBS", "Work Package", "Description"])
    for row in PHASE_SUMMARY:
        wbs.append(row)
    style_sheet(wbs)

    validation = wb.create_sheet("Validation Notes")
    validation.append(["Check", "Finding", "Correction"])
    validation_rows = [
        ("Required finish date", f"Original detailed Gantt ended on {original_issues['finish']}.", "Corrected baseline finishes on 2026-11-30, matching the report constraint."),
        ("Milestones", f"{len(original_issues['positive_duration_milestones'])} original milestone rows had positive durations.", "Corrected milestones use 0-day duration."),
        ("WBS workbook readability", "Original WBS workbook used a strict OOXML namespace that common parsers failed to expose as a normal sheet.", "Recovered source rows and created readable corrected CSV/XLSX outputs."),
        ("Duplicate wording", f"Duplicate titles found: {', '.join(original_issues['duplicate_task_titles'].keys()) or 'none'}.", "Corrected schedule uses unique activity names."),
    ]
    for row in validation_rows:
        validation.append(row)
    style_sheet(validation)

    wb.save(CORRECTED / "EMS_corrected_WBS_and_Gantt.xlsx")


def write_readme(schedule_rows: list[dict], original_issues: dict) -> None:
    start = schedule_rows[0]["Start"]
    finish = schedule_rows[-1]["Finish"]
    task_count = sum(1 for row in schedule_rows if row["Milestone"] == "No")
    milestone_count = sum(1 for row in schedule_rows if row["Milestone"] == "Yes")
    readme = f"""# EMS

Project Management case study for planning an Event Management System (EMS) for the 2027 Sydney Festival.

![Corrected EMS Gantt preview](assets/ems_gantt_preview.png)

## Scenario

Nouveau Event Creations (NEC), a subsidiary of Recreation Amalgamated Holdings (RAH), needs an integrated EMS to operate the 2027 Sydney Festival. In this academic scenario, Project Management Professionals Pty Ltd (PMP) plans the project for NEC. The EMS scope integrates registration, ticketing, performance scheduling, budget analysis, security management, vendor coordination, and operational reporting.

This repository presents the project as a project-management case study. It does not claim that the EMS was commissioned, built, or deployed for a real client.

## My Work

I worked from the PMP project-planning perspective and focused on turning the scenario into a coherent delivery plan:

- Defined the project scope around registration, ticketing, artist scheduling, budget analytics, security monitoring, vendor workflows, and NEC administrative reporting.
- Coordinated the WBS, AON dependency logic, Gantt baseline, milestones, risks, budget assumptions, resource roles, communication approach, and quality controls.
- Reviewed the original project outputs and corrected schedule/WBS inconsistencies before making the portfolio version public.
- Rebuilt the public-facing schedule so the plan aligns with the report constraint: EMS readiness by 30 November 2026, ahead of the 2027 Sydney Festival.
- Reviewed the original artifacts locally for traceability while adding corrected, recruiter-readable outputs.

## Corrected Planning Baseline

| Item | Value |
| --- | --- |
| Project | Event Management System for NEC |
| Case-study role | Project Management Professionals Pty Ltd (PMP) |
| Delivery window | {start} to {finish} |
| Required readiness date | 2026-11-30 |
| Corrected activity rows | {task_count} tasks and {milestone_count} milestones |
| Budget constraint | AUD 1.5 million |
| Primary deliverables | WBS, AON, Gantt baseline, risk/quality/change planning, stakeholder and communication planning |

## AON Logic

```mermaid
flowchart LR
  A[NEC objectives and MOV] --> B[Charter, scope, WBS, controls]
  B --> C[Architecture and security design]
  C --> D[Registration module]
  D --> E[Ticketing and QR/RFID validation]
  C --> F[Performance schedule module]
  C --> G[Budget analytics dashboard]
  C --> H[Security management dashboard]
  E --> I[Integrated EMS platform]
  F --> I
  G --> I
  H --> I
  I --> J[System, security, performance testing]
  J --> K[NEC UAT and operations rehearsal]
  K --> L[Production deployment and staff training]
  L --> M[Final contingency window]
  M --> N[EMS ready for 2027 Sydney Festival]
```

## Artifact Map

| Path | Purpose |
| --- | --- |
| `corrected/EMS_corrected_WBS_and_Gantt.xlsx` | Clean corrected workbook containing the WBS summary, corrected schedule, and validation notes. |
| `corrected/EMS_corrected_schedule.csv` | Recruiter-readable corrected Gantt source data. |
| `corrected/EMS_recovered_original_WBS.csv` | Recovered rows from the original WBS workbook for traceability. |
| `corrected/EMS_validation_report.md` | Detailed validation notes explaining what was fixed and why. |
| `assets/ems_gantt_preview.png` | Portfolio/GitHub visual showing the corrected phase-level Gantt. |
| `original-artifacts/` | Local-only source review folder. It is intentionally excluded from GitHub because the original report may contain assignment coversheet and jointly authored student details. |

## Validation Notes

- Original report states the EMS should be ready by 30 November 2026, but the original detailed Gantt extended to {original_issues['finish']}. The corrected baseline finishes on 2026-11-30.
- Original milestone rows used positive durations. The corrected schedule uses 0-day milestone rows.
- Original WBS workbook stored data in the package, but common workbook readers could not expose the sheet because of strict OOXML metadata. The recovered CSV and corrected workbook make the WBS readable.
- Duplicate activity wording in the source schedule was rewritten into distinct deliverable-based activities.

## Skills Demonstrated

Project planning, WBS structuring, AON dependency analysis, Gantt scheduling, milestone governance, stakeholder planning, communication planning, risk management, quality management, resource planning, budget constraint management, Excel-based project controls, and case-study documentation.
"""
    (PROJECT / "README.md").write_text(readme, encoding="utf-8")
    PUBLIC_CASE_STUDY.write_text(readme, encoding="utf-8")


def write_validation_report(original_issues: dict, schedule_rows: list[dict]) -> None:
    content = f"""# EMS Validation Report

## Source Issues Found

- Original Gantt task count: {original_issues['task_count']}
- Original Gantt date range: {original_issues['start']} to {original_issues['finish']}
- Required readiness date from report: 2026-11-30
- Positive-duration milestones found: {len(original_issues['positive_duration_milestones'])}
- Duplicate source activity names: {', '.join(original_issues['duplicate_task_titles'].keys()) or 'None'}
- Unknown predecessor references: {', '.join(original_issues['unknown_predecessor_refs']) or 'None'}
- Unknown successor references: {', '.join(original_issues['unknown_successor_refs']) or 'None'}

## Corrections Applied

- Rebuilt a coherent baseline from 2025-07-15 to 2026-11-30.
- Converted milestones to 0-day sign-off points.
- Aligned the WBS and Gantt around EMS modules: registration, ticketing, performance scheduling, budget analytics, security management, integration, testing, deployment, and closeout.
- Added an explicit contingency and defect-remediation window before the operational readiness milestone.
- Preserved source artifacts locally so the corrected public case study remains auditable without exposing private assignment details.

## Corrected Finish

The corrected schedule ends at `{schedule_rows[-1]['Finish']}`.
"""
    (CORRECTED / "EMS_validation_report.md").write_text(content, encoding="utf-8")


def write_gantt_preview(schedule_rows: list[dict]) -> None:
    rows = [row for row in schedule_rows if row["Milestone"] == "No"]
    phase_bounds = []
    for phase in dict.fromkeys(row["Phase"] for row in rows):
        phase_rows = [row for row in rows if row["Phase"] == phase]
        start = min(datetime.fromisoformat(row["Start"]) for row in phase_rows)
        end = max(datetime.fromisoformat(row["Finish"]) for row in phase_rows)
        phase_bounds.append((phase, start, end))

    width, height = 1800, 900
    margin_left, margin_right, margin_top, margin_bottom = 310, 170, 150, 120
    plot_w = width - margin_left - margin_right
    plot_h = height - margin_top - margin_bottom
    image = Image.new("RGB", (width, height), "#FFFFFF")
    draw = ImageDraw.Draw(image)
    try:
        font_title = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial Bold.ttf", 30)
        font_regular = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 18)
        font_small = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 16)
        font_bar = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial Bold.ttf", 17)
    except OSError:
        font_title = ImageFont.load_default()
        font_regular = ImageFont.load_default()
        font_small = ImageFont.load_default()
        font_bar = ImageFont.load_default()
    colors = ["#26A69A", "#5C6BC0", "#FFA726", "#EF5350", "#66BB6A"]
    min_date = min(start for _, start, _ in phase_bounds)
    max_date = datetime(2026, 12, 15)
    total_days = (max_date - min_date).days

    def x_pos(date: datetime) -> int:
        return margin_left + int(((date - min_date).days / total_days) * plot_w)

    draw.rectangle((0, 0, width, height), fill="#FFFFFF")
    draw.text((margin_left, 44), "EMS Project Management Case Study", fill="#111827", font=font_title)
    draw.text((margin_left, 86), "Corrected Gantt baseline aligned to 30 Nov 2026 readiness date", fill="#475569", font=font_regular)

    tick_dates = [
        datetime(2025, 7, 15),
        datetime(2025, 10, 1),
        datetime(2026, 1, 1),
        datetime(2026, 4, 1),
        datetime(2026, 7, 1),
        datetime(2026, 10, 1),
        datetime(2026, 11, 30),
    ]
    for tick in tick_dates:
        x = x_pos(tick)
        draw.line((x, margin_top, x, height - margin_bottom), fill="#D9E2EC", width=2)
        draw.text((x - 34, height - margin_bottom + 24), tick.strftime("%b %Y"), fill="#334155", font=font_small)

    row_gap = plot_h // len(phase_bounds)
    for idx, (phase, start, end) in enumerate(phase_bounds):
        y = margin_top + idx * row_gap + 18
        bar_h = 58
        x1, x2 = x_pos(start), x_pos(end)
        draw.text((60, y + 18), f"{idx + 1}", fill="#334155", font=font_regular)
        draw.text((100, y + 18), phase, fill="#111827", font=font_regular)
        draw.rounded_rectangle((x1, y, x2, y + bar_h), radius=8, fill=colors[idx % len(colors)])
        label = f"{start:%d %b %Y} - {end:%d %b %Y}"
        if x2 - x1 < 260:
            draw.text((max(margin_left, x1 - 220), y + 20), label, fill="#111827", font=font_small)
        else:
            draw.text((x1 + 16, y + 19), label, fill="#FFFFFF", font=font_bar)

    deadline = datetime(2026, 11, 30)
    x = x_pos(deadline)
    draw.line((x, margin_top - 20, x, height - margin_bottom), fill="#111827", width=4)
    draw.text((x - 230, margin_top - 56), "EMS ready: 30 Nov 2026", fill="#111827", font=font_regular)

    output = ASSETS / "ems_gantt_preview.png"
    image.save(output)
    image.save(PUBLIC_IMAGE)


def main() -> None:
    ensure_dirs()
    copy_originals()
    _, original_issues = read_original_gantt()
    recovered_wbs = recover_original_wbs_rows()
    schedule_rows = compute_schedule()
    write_csv(CORRECTED / "EMS_corrected_schedule.csv", schedule_rows)
    write_csv(CORRECTED / "EMS_recovered_original_WBS.csv", recovered_wbs)
    write_corrected_workbook(schedule_rows, original_issues)
    write_gantt_preview(schedule_rows)
    write_validation_report(original_issues, schedule_rows)
    write_readme(schedule_rows, original_issues)
    print(f"Wrote EMS project package to {PROJECT}")
    print(f"Corrected finish date: {schedule_rows[-1]['Finish']}")


if __name__ == "__main__":
    main()
